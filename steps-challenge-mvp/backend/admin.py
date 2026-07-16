from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List
import io
import requests as http_requests
import models
import schemas
from database import get_db
from auth import get_current_user

router = APIRouter(prefix="/api/admin", tags=["admin"])


def require_admin(current_user: models.User = Depends(get_current_user)):
    """Dependency to check if user is admin"""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user


@router.get("/users", response_model=List[schemas.AdminUserResponse])
def get_all_users(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Get all users with their total steps (admin only)"""
    users = db.query(
        models.User,
        func.coalesce(func.sum(models.StepEntry.steps), 0).label("total_steps")
    ).outerjoin(models.StepEntry).group_by(models.User.id).all()

    return [
        {
            "id": user.id,
            "full_name": user.full_name,
            "email": user.email,
            "team_name": user.team_name,
            "is_admin": user.is_admin,
            "is_approved": user.is_approved,
            "created_at": user.created_at,
            "total_steps": total_steps
        }
        for user, total_steps in users
    ]


@router.post("/users/{user_id}/approve")
def approve_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Approve a user registration (admin only)"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    user.is_approved = True
    db.commit()
    return {"message": f"User {user.full_name} approved successfully"}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Delete a user (admin only)"""
    if user_id == admin.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own admin account"
        )
    
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    db.delete(user)
    db.commit()
    return {"message": f"User {user.full_name} deleted successfully"}


@router.get("/steps", response_model=List[schemas.StepEntryResponse])
def get_all_steps(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Get all step entries (admin only)"""
    entries = db.query(models.StepEntry).order_by(models.StepEntry.date.desc()).all()
    return entries


@router.get("/users/{user_id}/steps", response_model=List[schemas.StepEntryResponse])
def get_user_steps(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Get all step entries for a specific user (admin only)"""
    # Verify user exists
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    entries = db.query(models.StepEntry).filter(
        models.StepEntry.user_id == user_id
    ).order_by(models.StepEntry.date.desc()).all()
    return entries


@router.delete("/steps/{entry_id}")
def delete_step_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Delete a step entry (admin only)"""
    entry = db.query(models.StepEntry).filter(models.StepEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Step entry not found"
        )
    
    db.delete(entry)
    db.commit()
    return {"message": "Step entry deleted successfully"}


@router.get("/stats")
def get_admin_stats(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Get overall statistics (admin only)"""
    total_users = db.query(func.count(models.User.id)).scalar()
    total_steps = db.query(func.sum(models.StepEntry.steps)).scalar() or 0
    total_entries = db.query(func.count(models.StepEntry.id)).scalar()
    
    return {
        "total_users": total_users,
        "total_steps": total_steps,
        "total_entries": total_entries,
        "teams": [
            "Overarching", "Release and DevOps", "Notifications",
            "Cloud Hosting", "Shared Services", "ECNS", "PAWHP",
            "Data", "CSP Legacy", "Not Listed",
        ]
    }


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Export all users and step entries as an Excel workbook (admin only)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Users Summary ──────────────────────────────────────────────
    ws_users = wb.active
    ws_users.title = "Users"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4B3B8C")
    center = Alignment(horizontal="center")

    user_headers = ["ID", "Full Name", "Email", "Team", "Total Steps", "Role", "Status", "Registered"]
    for col_idx, header in enumerate(user_headers, start=1):
        cell = ws_users.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    users_data = db.query(
        models.User,
        func.coalesce(func.sum(models.StepEntry.steps), 0).label("total_steps")
    ).outerjoin(models.StepEntry).group_by(models.User.id).order_by(
        func.coalesce(func.sum(models.StepEntry.steps), 0).desc()
    ).all()

    for row_idx, (user, total_steps) in enumerate(users_data, start=2):
        ws_users.cell(row=row_idx, column=1, value=user.id)
        ws_users.cell(row=row_idx, column=2, value=user.full_name)
        ws_users.cell(row=row_idx, column=3, value=user.email)
        ws_users.cell(row=row_idx, column=4, value=user.team_name)
        ws_users.cell(row=row_idx, column=5, value=int(total_steps))
        ws_users.cell(row=row_idx, column=6, value="Admin" if user.is_admin else "User")
        ws_users.cell(row=row_idx, column=7, value="Approved" if (user.is_approved or user.is_admin) else "Pending")
        ws_users.cell(row=row_idx, column=8, value=user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "")

    # Auto-size columns
    col_widths = [6, 28, 35, 22, 14, 8, 10, 18]
    for i, width in enumerate(col_widths, start=1):
        ws_users.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 2: All Step Entries ───────────────────────────────────────────
    ws_steps = wb.create_sheet(title="Step Entries")

    step_headers = ["Entry ID", "User ID", "Full Name", "Team", "Date", "Steps", "Logged At"]
    for col_idx, header in enumerate(step_headers, start=1):
        cell = ws_steps.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    entries = (
        db.query(models.StepEntry, models.User)
        .join(models.User, models.StepEntry.user_id == models.User.id)
        .order_by(models.StepEntry.date.desc())
        .all()
    )

    for row_idx, (entry, user) in enumerate(entries, start=2):
        ws_steps.cell(row=row_idx, column=1, value=entry.id)
        ws_steps.cell(row=row_idx, column=2, value=user.id)
        ws_steps.cell(row=row_idx, column=3, value=user.full_name)
        ws_steps.cell(row=row_idx, column=4, value=user.team_name)
        ws_steps.cell(row=row_idx, column=5, value=entry.date.strftime("%Y-%m-%d") if entry.date else "")
        ws_steps.cell(row=row_idx, column=6, value=entry.steps)
        ws_steps.cell(row=row_idx, column=7, value=entry.created_at.strftime("%Y-%m-%d %H:%M") if entry.created_at else "")

    step_col_widths = [10, 8, 28, 22, 12, 12, 18]
    for i, width in enumerate(step_col_widths, start=1):
        ws_steps.column_dimensions[get_column_letter(i)].width = width

    # ── Stream response ─────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=steps_challenge_export.xlsx"}
    )

@router.get("/export/photos-pdf")
def export_photos_pdf(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """Export all photos as a PDF (one photo per page, caption + uploader below) — admin only"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    photos = db.query(models.PhotoEntry, models.User).join(
        models.User, models.PhotoEntry.user_id == models.User.id
    ).order_by(models.PhotoEntry.created_at.asc()).all()

    if not photos:
        raise HTTPException(status_code=404, detail="No photos found")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    caption_style = ParagraphStyle(
        "caption",
        parent=styles["Normal"],
        fontSize=13,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1f2328"),
        fontName="Helvetica-Oblique",
    )
    uploader_style = ParagraphStyle(
        "uploader",
        parent=styles["Normal"],
        fontSize=11,
        leading=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4B3B8C"),
        fontName="Helvetica-Bold",
    )

    page_w = A4[0] - 3 * cm   # usable width
    page_h = A4[1] - 6 * cm   # usable height (leave room for text)
    max_img_h = page_h - 3 * cm

    story = []
    for idx, (photo, user) in enumerate(photos):
        # Fetch image bytes
        try:
            resp = http_requests.get(photo.image_url, timeout=15)
            resp.raise_for_status()
        except Exception:
            continue

        img_buf = io.BytesIO(resp.content)
        try:
            img = Image(img_buf)
        except Exception:
            continue

        # Scale to fit page while preserving aspect ratio
        iw, ih = img.imageWidth, img.imageHeight
        scale = min(page_w / iw, max_img_h / ih, 1.0)
        img.drawWidth = iw * scale
        img.drawHeight = ih * scale
        img._hAlign = "CENTER"

        story.append(img)
        story.append(Spacer(1, 0.4 * cm))

        if photo.caption:
            story.append(Paragraph(f'"{photo.caption}"', caption_style))
            story.append(Spacer(1, 0.2 * cm))

        story.append(Paragraph(f"Uploaded by {user.full_name}", uploader_style))

        if idx < len(photos) - 1:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=steps_challenge_photos.pdf"}
    )

# Made with Bob